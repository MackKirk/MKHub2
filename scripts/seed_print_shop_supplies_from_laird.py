"""Seed print shop supply products from Laird_Products.xlsx."""
from __future__ import annotations

import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.db import SessionLocal  # noqa: E402
from app.models.models import PrintShopSupplyProduct  # noqa: E402

DEFAULT_XLSX = Path(r"c:\Users\Raphael Coelho\Desktop\Laird_Products.xlsx")

CATEGORIES = [
    "Printing Rolls",
    "Laminating Rolls",
    "Boards",
    "Aplication Tape",
    "Application Tape",
    "Ink",
    "Other",
]


def _clean(s: str) -> str:
    s = (s or "").replace("\xa0", " ").replace("\u2002", " ")
    s = re.sub(r"^1x\s+", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _norm_category(name: str) -> str:
    if name == "Aplication Tape":
        return "Application Tape"
    return name


def _price_note(raw) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return str(raw)
    s = _clean(str(raw))
    return s or None


def parse_rows(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    category = "Other"
    sort = 0
    out = []
    for row in ws.iter_rows(values_only=True):
        name_raw = row[1] if len(row) > 1 else None
        price_raw = row[2] if len(row) > 2 else None
        note_raw = row[4] if len(row) > 4 else None
        if not name_raw:
            continue
        name = _clean(str(name_raw))
        if not name:
            continue
        if name in CATEGORIES:
            category = _norm_category(name)
            continue

        notes = _clean(str(note_raw)) if note_raw else None
        out.append(
            {
                "name": name,
                "category": category,
                "list_price_note": _price_note(price_raw),
                "notes": notes or None,
                "sort_index": sort,
            }
        )
        sort += 1
    wb.close()
    return out


def main():
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not path.exists():
        print(f"File not found: {path}")
        sys.exit(1)

    products = parse_rows(path)
    print(f"Parsed {len(products)} products from {path}")

    db = SessionLocal()
    try:
        # Ensure table exists
        from app.db import Base, engine

        Base.metadata.create_all(bind=engine, tables=[PrintShopSupplyProduct.__table__])

        existing = {r.name: r for r in db.query(PrintShopSupplyProduct).all()}
        now = datetime.now(timezone.utc)
        created = 0
        updated = 0
        for p in products:
            row = existing.get(p["name"])
            if row:
                row.category = p["category"]
                row.list_price_note = p["list_price_note"]
                if p["notes"]:
                    row.notes = p["notes"]
                row.sort_index = p["sort_index"]
                row.is_active = True
                row.updated_at = now
                updated += 1
            else:
                db.add(
                    PrintShopSupplyProduct(
                        name=p["name"],
                        category=p["category"],
                        unit="ea",
                        list_price_note=p["list_price_note"],
                        notes=p["notes"],
                        stock_quantity=0,
                        reorder_point=0,
                        sort_index=p["sort_index"],
                        is_active=True,
                        created_at=now,
                        updated_at=now,
                    )
                )
                created += 1
        db.commit()
        print(f"Done. created={created} updated={updated}")
        cats = db.execute(
            text(
                "SELECT category, COUNT(*) FROM print_shop_supply_products GROUP BY category ORDER BY category"
            )
        ).fetchall()
        for c, n in cats:
            print(f"  {c}: {n}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
