"""
Script to import fleet heavy machinery from the Equip. Master List tab of an Excel file.

Maps columns from the Excel sheet to FleetAsset model fields.
Creates assets with asset_type='heavy_machinery'.

Usage:
    python scripts/import_fleet_heavy_machinery_from_excel.py <path_to_xlsx> [--dry-run]

Expected columns (Equip. Master List tab):
    UNIT #, PLATE #, VIN #, YEAR, MAKE, MODEL, TYPE, GVW, CVIP/NDT EXPIRY, NOTES
"""
import sys
import os
import re

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Load environment variables first
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception as e:
    print(f"WARNING: Could not load .env file: {e}")

# Check database type before importing
database_url = os.getenv("DATABASE_URL", "sqlite:///./var/dev.db")
if database_url.startswith("postgresql"):
    try:
        import psycopg2
    except ImportError:
        print("ERROR: PostgreSQL database detected but psycopg2 is not installed.")
        sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

try:
    from app.db import SessionLocal
    from app.models.models import FleetAsset
except ImportError as e:
    print(f"ERROR: Failed to import database components: {e}")
    sys.exit(1)

# Column name aliases for Equip. Master List
COLUMN_ALIASES = {
    "UNIT #": ["UNIT #", "UNIT#", "UNIT", "UNIT NUMBER"],
    "PLATE #": ["PLATE #", "PLATE#", "PLATE", "LICENSE PLATE"],
    "VIN #": ["VIN #", "VIN#", "VIN", "SERIAL"],
    "YEAR": ["YEAR"],
    "MAKE": ["MAKE"],
    "MODEL": ["MODEL", "MODEL "],
    "TYPE": ["TYPE", "EQUIPMENT TYPE", "EQUIP TYPE"],
    "GVW": ["GVW", "GVWR", "GVW (KG)"],
    "CVIP/NDT EXPIRY": ["CVIP/NDT EXPIRY", "CVIP/NDT", "CVIP NDT EXPIRY", "CVIP EXPIRY"],
    "NOTES": ["NOTES"],
}


def normalize_field(value) -> str | None:
    """Normalize a string field, strip whitespace."""
    if value is None or (isinstance(value, float) and str(value) == "nan"):
        return None
    value = str(value).strip()
    return value if value else None


def get_cell_value(cell) -> str | None:
    """Get cell value as string, handling None and empty."""
    if cell is None or cell.value is None:
        return None
    val = cell.value
    if isinstance(val, float) and (val != val or str(val) == "nan"):
        return None
    return str(val).strip() or None


def parse_year(val) -> int | None:
    """Parse year to integer."""
    if val is None or val == "":
        return None
    try:
        if isinstance(val, (int, float)):
            y = int(val)
            if 1900 <= y <= 2100:
                return y
        s = str(val).strip()
        if not s:
            return None
        y = int(re.sub(r"[^\d]", "", s)[:4])
        if 1900 <= y <= 2100:
            return y
    except (ValueError, TypeError):
        pass
    return None


def parse_gvwr(val) -> int | None:
    """Parse GVW to integer (kg). Handles lbs suffix - converts to kg."""
    if val is None or val == "":
        return None
    try:
        s = str(val).strip().upper()
        is_lbs = "LBS" in s or "LB" in s
        num_str = re.sub(r"[^\d.]", "", s)
        if not num_str:
            return None
        num = float(num_str)
        if is_lbs:
            num = num * 0.453592
        return int(round(num))
    except (ValueError, TypeError):
        return None


def find_column_index(headers: list[str], canonical: str) -> int | None:
    """Find column index by header name, using aliases."""
    aliases = COLUMN_ALIASES.get(canonical, [canonical])
    key_parts = []
    for a in aliases:
        a_clean = (a or "").strip().upper()
        key_parts.append(a_clean)
        first_word = a_clean.split()[0] if a_clean else ""
        if first_word and len(first_word) >= 2:
            key_parts.append(first_word)

    for i, h in enumerate(headers):
        h_clean = (h or "").strip().upper()
        if not h_clean:
            continue
        for term in key_parts:
            if h_clean == term or term in h_clean:
                return i
    return None


def is_row_empty(row: list) -> bool:
    """Check if row has any non-empty cell."""
    for cell in row:
        val = cell.value if hasattr(cell, "value") else cell
        if val is not None and str(val).strip():
            return False
    return True


def import_fleet_heavy_machinery(excel_path: str, dry_run: bool = False):
    """Import fleet heavy machinery from Equip. Master List sheet."""

    if not os.path.exists(excel_path):
        print(f"ERROR: File not found: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheet_name = "Equip. Master List"
    ws = None
    for sn in wb.sheetnames:
        if sn.strip() == sheet_name or "Equip" in sn and "Master" in sn:
            ws = wb[sn]
            break
    if ws is None:
        print(f"ERROR: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        sys.exit(1)
    rows = list(ws.iter_rows())
    wb.close()

    if len(rows) < 2:
        print("ERROR: Sheet has no data rows (header + at least one row required).")
        sys.exit(1)

    # Find header row
    header_row_idx = 0
    for idx, row in enumerate(rows[:10]):
        vals = [str(c.value or "").strip().upper() for c in row]
        if any("UNIT" in v or "MAKE" in v or "VIN" in v or "TYPE" in v for v in vals if v):
            header_row_idx = idx
            break

    header_row = rows[header_row_idx]
    headers = [str(c.value or "").strip() for c in header_row]

    col_map = {}
    for canonical in [
        "UNIT #", "PLATE #", "VIN #", "YEAR", "MAKE", "MODEL",
        "TYPE", "GVW", "CVIP/NDT EXPIRY", "NOTES",
    ]:
        idx = find_column_index(headers, canonical)
        if idx is not None:
            col_map[canonical] = idx

    print(f"Columns found: {col_map}")
    print(f"\n{'[DRY RUN] ' if dry_run else ''}Starting import...\n")

    db = SessionLocal()
    created_count = 0
    skipped_count = 0
    error_count = 0

    data_rows = rows[header_row_idx + 1 :]

    try:
        for row_num, row in enumerate(data_rows, start=header_row_idx + 2):
            if is_row_empty(row):
                skipped_count += 1
                continue

            try:
                def get(col: str):
                    idx = col_map.get(col)
                    if idx is None or idx >= len(row):
                        return None
                    return get_cell_value(row[idx])

                unit_number = normalize_field(get("UNIT #"))
                make = normalize_field(get("MAKE"))
                model = normalize_field(get("MODEL"))
                vin = normalize_field(get("VIN #") or get("VIN"))

                name = None
                if make and model:
                    name = f"{make} {model}".strip()
                elif make:
                    name = make
                elif model:
                    name = model
                elif unit_number:
                    name = f"Equipment {unit_number}"
                elif vin:
                    name = vin
                else:
                    print(f"Row {row_num}: SKIP - no UNIT #, MAKE/MODEL, or VIN #")
                    skipped_count += 1
                    continue

                # Duplicate check
                if unit_number:
                    existing = db.query(FleetAsset).filter(
                        FleetAsset.unit_number == unit_number,
                        FleetAsset.asset_type == "heavy_machinery",
                    ).first()
                    if existing:
                        print(f"Row {row_num}: SKIP - unit_number '{unit_number}' already exists")
                        skipped_count += 1
                        continue

                if vin:
                    existing = db.query(FleetAsset).filter(
                        FleetAsset.vin == vin,
                        FleetAsset.asset_type == "heavy_machinery",
                    ).first()
                    if existing:
                        print(f"Row {row_num}: SKIP - VIN '{vin}' already exists")
                        skipped_count += 1
                        continue

                # Build notes: combine CVIP/NDT EXPIRY and NOTES
                cvip_expiry = normalize_field(get("CVIP/NDT EXPIRY"))
                notes_raw = normalize_field(get("NOTES"))
                notes_parts = []
                if cvip_expiry:
                    notes_parts.append(f"CVIP/NDT expiry: {cvip_expiry}")
                if notes_raw:
                    notes_parts.append(notes_raw)
                notes = "\n".join(notes_parts) if notes_parts else None

                asset_data = {
                    "asset_type": "heavy_machinery",
                    "name": name,
                    "unit_number": unit_number,
                    "vin": vin,
                    "license_plate": normalize_field(get("PLATE #")),
                    "make": make,
                    "model": model,
                    "year": parse_year(get("YEAR")),
                    "equipment_type_label": normalize_field(get("TYPE")),
                    "gvw_kg": parse_gvwr(get("GVW")),
                    "notes": notes,
                    "status": "active",
                }

                final_data = {}
                for k, v in asset_data.items():
                    if v is None:
                        continue
                    if isinstance(v, str) and not v.strip():
                        continue
                    final_data[k] = v

                if not dry_run:
                    asset = FleetAsset(**final_data)
                    db.add(asset)
                    db.commit()
                    db.refresh(asset)
                    created_count += 1
                    print(f"Row {row_num}: OK - created '{name}' (unit={unit_number})")
                else:
                    created_count += 1
                    print(f"Row {row_num}: [DRY RUN] would create '{name}' (unit={unit_number})")

            except Exception as e:
                error_count += 1
                print(f"Row {row_num}: ERROR - {e}")
                if not dry_run:
                    db.rollback()

        print(f"\n{'='*60}")
        print("Import complete!")
        print(f"  Created: {created_count}")
        print(f"  Skipped: {skipped_count}")
        print(f"  Errors:  {error_count}")
        print(f"{'='*60}")

    except Exception as e:
        print(f"ERROR: {e}")
        if not dry_run:
            db.rollback()
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python scripts/import_fleet_heavy_machinery_from_excel.py <path_to_xlsx> [--dry-run]")
        sys.exit(1)

    excel_path = sys.argv[1]
    dry_run = "--dry-run" in sys.argv or "-d" in sys.argv

    import_fleet_heavy_machinery(excel_path, dry_run=dry_run)
