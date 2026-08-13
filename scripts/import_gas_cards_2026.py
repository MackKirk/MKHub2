"""
Import 2026 gas cards from the Gas Card Information workbook into fuel_cards.

Usage:
  python scripts/import_gas_cards_2026.py
  python scripts/import_gas_cards_2026.py --path "C:/path/to/2026 - Gas Card Information.xlsx"
  python scripts/import_gas_cards_2026.py --dry-run
"""
from __future__ import annotations

import argparse
import re
import sys
import os
from datetime import date, datetime, timezone
from typing import Optional

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from dotenv import load_dotenv

    load_dotenv()
except Exception:
    pass

from openpyxl import load_workbook
from sqlalchemy import text

from app.db import SessionLocal, engine, Base
from app.models.models import EmployeeProfile, FuelCard, FuelCardAssignment


DEFAULT_PATH = r"c:\Users\Raphael Coelho\Desktop\2026 - Gas Card Information.xlsx"
SHEET_NAME = "2026 - Gas Card"


def _norm_space(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def _norm_key(s: str) -> str:
    s = _norm_space(s).lower()
    s = re.sub(r"[^\w\s]", " ", s)
    return _norm_space(s)


def normalize_crew(raw) -> Optional[str]:
    if raw is None:
        return None
    s = _norm_space(str(raw))
    if not s or s in {"-", "—", "n/a", "na"}:
        return None
    if s.upper() == "CW":
        return "C & W"
    return s


def parse_issue_date(raw) -> Optional[date]:
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = _norm_space(str(raw))
    if not s:
        return None
    # e.g. 2026-Feb-10 / 2026-MAY-05 / 2026-Mar-31
    for fmt in ("%Y-%b-%d", "%Y-%B-%d", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # case-insensitive month abbreviations
    try:
        parts = s.replace("/", "-").split("-")
        if len(parts) == 3:
            y, mon, d = parts
            months = {
                "jan": 1,
                "feb": 2,
                "mar": 3,
                "apr": 4,
                "may": 5,
                "jun": 6,
                "jul": 7,
                "aug": 8,
                "sep": 9,
                "oct": 10,
                "nov": 11,
                "dec": 12,
            }
            m = months.get(mon[:3].lower())
            if m:
                return date(int(y), m, int(d))
    except Exception:
        pass
    print(f"  ! could not parse date: {raw!r}")
    return None


def clean_holder_name(raw) -> Optional[str]:
    if raw is None:
        return None
    s = _norm_space(str(raw))
    if not s:
        return None
    upper = s.upper()
    if upper in {"CANCELLED", "CANCELED", "LOST", "RETURNED", "N/A", "-"}:
        return None
    # "C & W - Matt Ius" / "C&W - Matt Ius"
    s = re.sub(r"^(?:c\s*&\s*w|cw)\s*[-–—:]\s*", "", s, flags=re.IGNORECASE)
    # drop parenthetical nicknames for matching variants: keep both full and without
    return _norm_space(s) or None


def name_variants(full: str) -> list[str]:
    """Return normalized match keys for a display name."""
    full = _norm_space(full)
    keys = set()
    keys.add(_norm_key(full))
    # remove parenthetical nicknames
    no_paren = _norm_space(re.sub(r"\([^)]*\)", " ", full))
    if no_paren:
        keys.add(_norm_key(no_paren))
    # nickname only from parentheses
    m = re.search(r"\(([^)]+)\)", full)
    if m and no_paren:
        nick = _norm_space(m.group(1))
        parts = no_paren.split()
        if len(parts) >= 2 and nick:
            keys.add(_norm_key(f"{nick} {parts[-1]}"))
    return [k for k in keys if k]


def build_employee_index(db):
    """Map normalized full-name keys -> list of (user_id, display). Also last-name buckets."""
    index: dict[str, list[tuple]] = {}
    by_last: dict[str, list[tuple]] = {}
    rows = db.query(EmployeeProfile).filter(EmployeeProfile.user_id.isnot(None)).all()
    for e in rows:
        first = _norm_space(e.first_name or "")
        last = _norm_space(e.last_name or "")
        preferred = _norm_space(e.preferred_name or "")
        candidates = []
        if first and last:
            candidates.append(f"{first} {last}")
        if preferred and last:
            candidates.append(f"{preferred} {last}")
            # preferred may already be a full name ("Matt Stoesz")
            if " " in preferred:
                candidates.append(preferred)
        if preferred and not last:
            candidates.append(preferred)
        if preferred and " " not in preferred:
            candidates.append(preferred)  # nickname alone (e.g. Sonny)
        if preferred and first and last and preferred.lower() != first.lower() and " " not in preferred:
            candidates.append(f"{preferred} {last}")
        display = preferred if preferred and " " not in preferred else (preferred or first)
        if preferred and " " in preferred:
            display = preferred
        else:
            display = f"{(preferred or first)} {last}".strip() if last else (preferred or first)
        entry = (e.user_id, display or f"{first} {last}".strip(), first, last, preferred)
        for c in candidates:
            for key in name_variants(c):
                index.setdefault(key, []).append(entry)
        if last:
            by_last.setdefault(_norm_key(last), []).append(entry)
    return index, by_last


def _first_tokens_compatible(a: str, b: str) -> bool:
    a = _norm_key(a)
    b = _norm_key(b)
    if not a or not b:
        return False
    if a == b:
        return True
    # Matt / Matthew, Sam / Samuel, Dan / Daniel, Dave / David, Jay / Jason
    shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
    if len(shorter) >= 3 and longer.startswith(shorter):
        return True
    # token overlap ("Masum" in "MD Masum")
    a_parts = set(a.split())
    b_parts = set(b.split())
    return bool(a_parts & b_parts)


def resolve_employee(index, by_last, holder_name: Optional[str]):
    if not holder_name:
        return None, "no_name"

    # Exact / variant keys
    for key in name_variants(holder_name):
        hits = index.get(key) or []
        uniq = {}
        for uid, disp, *_rest in hits:
            uniq[str(uid)] = (uid, disp)
        if len(uniq) == 1:
            uid, disp = next(iter(uniq.values()))
            return (uid, disp), "matched"
        if len(uniq) > 1:
            return None, f"ambiguous:{', '.join(d for _, d in uniq.values())}"

    # Nickname in parentheses against preferred-only index
    nick_m = re.search(r"\(([^)]+)\)", holder_name)
    if nick_m:
        nick = _norm_space(nick_m.group(1))
        hits = index.get(_norm_key(nick)) or []
        uniq = {}
        for uid, disp, *_rest in hits:
            uniq[str(uid)] = (uid, disp)
        if len(uniq) == 1:
            uid, disp = next(iter(uniq.values()))
            return (uid, disp), "matched_nick"

    # Last-name unique + compatible first/preferred
    cleaned = _norm_space(re.sub(r"\([^)]*\)", " ", holder_name))
    parts = cleaned.split()
    if len(parts) >= 2:
        last_key = _norm_key(parts[-1])
        first_token = parts[0]
        # also try first two tokens as first name ("MD Masum Billah" / "Duc Son")
        first_join = " ".join(parts[:-1])
        bucket = by_last.get(last_key) or []
        # typo-tolerant last names (Vbra/Vrba, Cumming/Cummings)
        if not bucket and len(last_key) >= 4:
            for lk, ents in by_last.items():
                if lk.startswith(last_key) or last_key.startswith(lk):
                    if abs(len(lk) - len(last_key)) <= 1:
                        bucket = ents
                        break
                if len(lk) == len(last_key):
                    diffs = [i for i in range(len(lk)) if lk[i] != last_key[i]]
                    if len(diffs) == 2 and lk[diffs[0]] == last_key[diffs[1]] and lk[diffs[1]] == last_key[diffs[0]]:
                        bucket = ents
                        break
                    if len(diffs) == 1:
                        bucket = ents
                        break
        candidates = []
        for uid, disp, first, last, preferred in bucket:
            pref_token = preferred.split()[0] if preferred else ""
            if (
                _first_tokens_compatible(first_token, first)
                or _first_tokens_compatible(first_token, pref_token)
                or _first_tokens_compatible(first_join, first)
                or (preferred and _norm_key(preferred) == _norm_key(cleaned))
                or (preferred and _norm_key(preferred) == _norm_key(holder_name))
            ):
                candidates.append((uid, disp))
        uniq = {str(uid): (uid, disp) for uid, disp in candidates}
        if len(uniq) == 1:
            uid, disp = next(iter(uniq.values()))
            return (uid, disp), "matched_fuzzy"
        if len(uniq) > 1:
            return None, f"ambiguous:{', '.join(d for _, d in uniq.values())}"

    return None, "not_found"


def infer_status(name_raw, pin_raw, crew_raw) -> str:
    blob = " ".join(str(x).upper() for x in (name_raw, pin_raw, crew_raw) if x is not None)
    if "LOST" in blob:
        return "lost"
    if "CANCEL" in blob or "DOES NOT WORK" in blob:
        return "cancelled"
    return "active"


def ensure_schema(db):
    Base.metadata.create_all(bind=engine, tables=[FuelCard.__table__, FuelCardAssignment.__table__])
    try:
        db.execute(text("ALTER TABLE fuel_cards ADD COLUMN IF NOT EXISTS crew VARCHAR(100)"))
        db.execute(text("ALTER TABLE fuel_cards ALTER COLUMN date_issued DROP NOT NULL"))
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"schema migrate note: {e}")


def read_rows(path: str):
    wb = load_workbook(path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet {SHEET_NAME!r} not found. Available: {wb.sheetnames}")
    ws = wb[SHEET_NAME]
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        if row is None or row[0] is None or str(row[0]).strip() == "":
            continue
        out.append(
            {
                "card_number": _norm_space(str(row[0])),
                "name_raw": row[1],
                "pin": _norm_space(str(row[2])) if row[2] is not None else "",
                "crew_raw": row[3],
                "date_raw": row[4],
            }
        )
    wb.close()
    return out


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--path", default=DEFAULT_PATH)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not os.path.isfile(args.path):
        raise SystemExit(f"File not found: {args.path}")

    rows = read_rows(args.path)
    print(f"Read {len(rows)} gas card rows from {args.path}")

    db = SessionLocal()
    try:
        ensure_schema(db)
        emp_index, by_last = build_employee_index(db)
        print(f"Employee name keys indexed: {len(emp_index)}")

        created = 0
        updated = 0
        assigned = 0
        skipped_assign = []
        now = datetime.now(timezone.utc)

        for row in rows:
            card_number = row["card_number"]
            pin = row["pin"] or "0000"
            crew = normalize_crew(row["crew_raw"])
            date_issued = parse_issue_date(row["date_raw"])
            status = infer_status(row["name_raw"], row["pin"], row["crew_raw"])
            holder = clean_holder_name(row["name_raw"])

            existing = db.query(FuelCard).filter(FuelCard.card_number == card_number).first()
            if existing:
                existing.pin = pin
                existing.crew = crew
                existing.date_issued = date_issued
                existing.status = status
                existing.updated_at = now
                card = existing
                updated += 1
                action = "update"
            else:
                card = FuelCard(
                    card_number=card_number,
                    pin=pin,
                    date_issued=date_issued,
                    crew=crew,
                    status=status,
                    notes=None,
                )
                db.add(card)
                db.flush()
                created += 1
                action = "create"

            match, reason = resolve_employee(emp_index, by_last, holder)
            assign_note = None
            if status != "active":
                assign_note = f"skip assign ({status})"
            elif not holder:
                assign_note = "skip assign (no holder name)"
            elif match is None:
                assign_note = f"skip assign ({reason}): {holder}"
                skipped_assign.append((card_number, holder, reason))
            else:
                user_id, display = match
                active = (
                    db.query(FuelCardAssignment)
                    .filter(
                        FuelCardAssignment.fuel_card_id == card.id,
                        FuelCardAssignment.is_active == True,  # noqa: E712
                    )
                    .first()
                )
                if active and active.assigned_to_user_id == user_id:
                    assign_note = f"already assigned -> {display}"
                else:
                    if active:
                        active.is_active = False
                        active.returned_at = now
                    db.add(
                        FuelCardAssignment(
                            fuel_card_id=card.id,
                            assigned_to_user_id=user_id,
                            assigned_at=now,
                            is_active=True,
                            notes="Imported from 2026 Gas Card Information",
                        )
                    )
                    assigned += 1
                    assign_note = f"assigned -> {display}"

            print(
                f"[{action}] #{card_number} pin={pin} crew={crew or '-'} "
                f"issued={date_issued or '-'} status={status} holder={holder or '-'} | {assign_note}"
            )

        if args.dry_run:
            db.rollback()
            print("\nDRY RUN — rolled back")
        else:
            db.commit()
            print("\nCommitted")

        print(
            f"Summary: created={created} updated={updated} assigned={assigned} "
            f"unmatched_holders={len(skipped_assign)}"
        )
        if skipped_assign:
            print("\nCould not match (assign later):")
            for card_number, holder, reason in skipped_assign:
                print(f"  #{card_number}  {holder}  ({reason})")
    finally:
        db.close()


if __name__ == "__main__":
    main()
